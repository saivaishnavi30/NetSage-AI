import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, Alignment

# --------------------------------------------------
# Paths
# --------------------------------------------------

CASES_FILE = "data/cases.csv"
REVIEW_FILE = "review/responsible_ai_log.csv"
OUTPUT_FILE = "outputs/NetSage_Dashboard.xlsx"

os.makedirs("outputs", exist_ok=True)

# --------------------------------------------------
# Load data
# --------------------------------------------------

cases = pd.read_csv(CASES_FILE)
review = pd.read_csv(REVIEW_FILE)

# --------------------------------------------------
# Summary calculations
# --------------------------------------------------

issue_summary = (
    cases["issue_type"]
    .value_counts()
    .reset_index()
)

issue_summary.columns = ["Issue Type", "Count"]

severity_summary = (
    cases["severity"]
    .value_counts()
    .reset_index()
)

severity_summary.columns = ["Severity", "Count"]

total_cases = len(cases)
reviewed_cases = len(review)

accepted_cases = (
    review["human_decision"]
    .astype(str)
    .str.lower()
    .eq("accepted")
    .sum()
)

edited_cases = (
    review["human_decision"]
    .astype(str)
    .str.lower()
    .eq("edited")
    .sum()
)

rejected_cases = (
    review["human_decision"]
    .astype(str)
    .str.lower()
    .eq("rejected")
    .sum()
)

agreement_rate = (
    accepted_cases / reviewed_cases * 100
    if reviewed_cases > 0
    else 0
)

# --------------------------------------------------
# Create Excel workbook
# --------------------------------------------------

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    cases.to_excel(
        writer,
        sheet_name="Cases",
        index=False
    )

    review.to_excel(
        writer,
        sheet_name="Responsible AI",
        index=False
    )

    issue_summary.to_excel(
        writer,
        sheet_name="Issue Summary",
        index=False
    )

    severity_summary.to_excel(
        writer,
        sheet_name="Severity Summary",
        index=False
    )

    dashboard_data = pd.DataFrame({
        "Metric": [
            "Total Cases",
            "Human Reviewed",
            "Accepted",
            "Edited",
            "Rejected",
            "AI-Human Agreement %"
        ],
        "Value": [
            total_cases,
            reviewed_cases,
            accepted_cases,
            edited_cases,
            rejected_cases,
            agreement_rate
        ]
    })

    dashboard_data.to_excel(
        writer,
        sheet_name="Dashboard",
        index=False
    )

# --------------------------------------------------
# Format workbook
# --------------------------------------------------

wb = load_workbook(OUTPUT_FILE)

# Dashboard
ws = wb["Dashboard"]

ws["D1"] = "NetSage AI"
ws["D1"].font = Font(
    size=20,
    bold=True
)

ws["D2"] = "Network Troubleshooting & Responsible AI Dashboard"
ws["D2"].font = Font(
    size=12,
    bold=True
)

# Make columns readable
for sheet in wb.worksheets:

    for column in sheet.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:

            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 45)

    for row in sheet.iter_rows():

        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

# --------------------------------------------------
# Issue Type Chart
# --------------------------------------------------

issue_ws = wb["Issue Summary"]

chart1 = BarChart()

chart1.title = "Cases by Issue Type"
chart1.y_axis.title = "Number of Cases"
chart1.x_axis.title = "Issue Type"

data = Reference(
    issue_ws,
    min_col=2,
    min_row=1,
    max_row=len(issue_summary) + 1
)

categories = Reference(
    issue_ws,
    min_col=1,
    min_row=2,
    max_row=len(issue_summary) + 1
)

chart1.add_data(
    data,
    titles_from_data=True
)

chart1.set_categories(categories)

ws.add_chart(chart1, "D5")

# --------------------------------------------------
# Severity Chart
# --------------------------------------------------

severity_ws = wb["Severity Summary"]

chart2 = PieChart()

chart2.title = "Cases by Severity"

data = Reference(
    severity_ws,
    min_col=2,
    min_row=1,
    max_row=len(severity_summary) + 1
)

labels = Reference(
    severity_ws,
    min_col=1,
    min_row=2,
    max_row=len(severity_summary) + 1
)

chart2.add_data(
    data,
    titles_from_data=True
)

chart2.set_categories(labels)

ws.add_chart(chart2, "D22")

# --------------------------------------------------
# Save
# --------------------------------------------------

wb.save(OUTPUT_FILE)

print("=" * 60)
print("NetSage AI Excel Dashboard")
print("=" * 60)
print(f"Total cases: {total_cases}")
print(f"Human reviewed: {reviewed_cases}")
print(f"Accepted: {accepted_cases}")
print(f"Edited: {edited_cases}")
print(f"Rejected: {rejected_cases}")
print(f"AI-Human agreement: {agreement_rate:.1f}%")
print("-" * 60)
print(f"Excel created: {OUTPUT_FILE}")
print("=" * 60)
